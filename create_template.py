# -*- coding: utf-8 -*-
"""
批量处理 Excel 模板生成脚本
运行此脚本可生成批量处理所需的 Excel 模板文件
"""
import openpyxl


def create_template():
    """创建批量处理 Excel 模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批量处理模板"

    # 设置表头
    headers = ["模式类型", "学校名称/ID", "账号", "密码", "userId", "备注"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="DDDDDD",
            end_color="DDDDDD",
            fill_type="solid"
        )

    # 添加示例数据
    # userId 模式示例
    ws.cell(row=2, column=1, value="userid")
    ws.cell(row=2, column=5, value="1234567890123456789")
    ws.cell(row=2, column=6, value="userId 模式示例")

    # 登录模式示例（学校名称）
    ws.cell(row=3, column=1, value="login")
    ws.cell(row=3, column=2, value="南京晓庄学院")
    ws.cell(row=3, column=3, value="student001")
    ws.cell(row=3, column=4, value="password123")
    ws.cell(row=3, column=6, value="登录模式示例（学校名称）")

    # 登录模式示例（学校ID）
    ws.cell(row=4, column=1, value="login")
    ws.cell(row=4, column=2, value="1224316234189443073")
    ws.cell(row=4, column=3, value="student002")
    ws.cell(row=4, column=4, value="password456")
    ws.cell(row=4, column=6, value="登录模式示例（学校ID）")

    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20

    # 保存文件
    filename = "批量处理模板.xlsx"
    wb.save(filename)
    print(f"模板文件已生成：{filename}")
    print("\n使用说明：")
    print("1. 打开模板文件，按照格式填写账号信息")
    print("2. 模式类型：填写 'login' 或 'userid'")
    print("3. userId 模式：仅需填写 userId 列（19位纯数字）")
    print("4. 登录模式：需要填写学校名称/ID、账号、密码")
    print("5. 在 GUI 中选择'批量处理'标签页，导入 Excel 文件即可")


if __name__ == "__main__":
    create_template()