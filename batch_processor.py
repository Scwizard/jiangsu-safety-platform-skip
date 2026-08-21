# -*- coding: utf-8 -*-
"""
批量化处理模块

支持通过 Excel 表格批量处理多个账号的 userId 模式或登录模式。
"""
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Dict, Callable, Optional
import threading
import queue

import engine
import userid_validator


class BatchTask:
    """单个批量任务"""
    def __init__(self, mode: str, school: str = "", account: str = "",
                 password: str = "", userid: str = "", remark: str = ""):
        self.mode = mode  # "login" 或 "userid"
        self.school = school
        self.account = account
        self.password = password
        self.userid = userid
        self.remark = remark


class TaskResult:
    """单个任务的处理结果"""
    def __init__(self, task: BatchTask, success: bool, score: int = 0,
                 error: str = "", duration: float = 0):
        self.task = task
        self.success = success
        self.score = score
        self.error = error
        self.duration = duration


class BatchResult:
    """批量处理结果汇总"""
    def __init__(self):
        self.total = 0
        self.success_count = 0
        self.fail_count = 0
        self.details: List[TaskResult] = []


class BatchProcessor:
    """批量处理器"""

    def __init__(self):
        self._cancel_event = threading.Event()
        self._log_queue = queue.Queue()

    def create_userid_template(self, save_path: str) -> str:
        """
        生成 userId 模式 Excel 模板

        Args:
            save_path: 保存路径

        Returns:
            保存的文件路径
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "userId批量处理模板"

        # 设置表头
        headers = ["userId", "备注"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # 添加示例数据
        ws.cell(row=2, column=1, value="1234567890123456789")
        ws.cell(row=2, column=2, value="学生A")

        ws.cell(row=3, column=1, value="9876543210987654321")
        ws.cell(row=3, column=2, value="学生B")

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

        wb.save(save_path)
        return save_path

    def create_login_template(self, save_path: str) -> str:
        """
        生成登录模式 Excel 模板

        Args:
            save_path: 保存路径

        Returns:
            保存的文件路径
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "登录批量处理模板"

        # 设置表头
        headers = ["学校名称/ID", "账号", "密码", "备注"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # 添加示例数据
        ws.cell(row=2, column=1, value="南京晓庄学院")
        ws.cell(row=2, column=2, value="student001")
        ws.cell(row=2, column=3, value="password123")
        ws.cell(row=2, column=4, value="学生C")

        ws.cell(row=3, column=1, value="1224316234189443073")
        ws.cell(row=3, column=2, value="student002")
        ws.cell(row=3, column=3, value="password456")
        ws.cell(row=3, column=4, value="学生D")

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        wb.save(save_path)
        return save_path

    def parse_excel_by_mode(self, mode: str, file_path: str) -> tuple[List[BatchTask], str]:
        """
        根据指定模式解析 Excel 文件

        Args:
            mode: "userid" 或 "login"
            file_path: Excel 文件路径

        Returns:
            (任务列表, 错误信息) - 如果解析成功，错误信息为空字符串
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            tasks = []
            errors = []

            # 假设第一行是表头，从第二行开始读取
            header_row = 1
            headers = [cell.value for cell in sheet[header_row]]

            if mode == "userid":
                # userId 模式：需要 userId 列
                if "userId" not in headers:
                    return [], "Excel 文件缺少必需的列：userId"

                userid_idx = headers.index("userId")
                remark_idx = headers.index("备注") if "备注" in headers else -1

                for row_idx in range(2, sheet.max_row + 1):
                    row = sheet[row_idx]
                    userid = row[userid_idx].value

                    if not userid:
                        continue  # 跳过空行

                    userid = str(userid).strip()

                    # 使用统一的校验模块验证 userId 格式
                    is_valid, result = userid_validator.validate_and_clean_userid(userid)
                    if not is_valid:
                        errors.append(f"第 {row_idx} 行：{result}")
                        continue
                    userid = result  # 使用清理后的 userId

                    remark = row[remark_idx].value if remark_idx >= 0 else ""
                    remark = str(remark).strip() if remark else ""

                    task = BatchTask(mode="userid", userid=userid, remark=remark)
                    tasks.append(task)

            elif mode == "login":
                # 登录模式：需要学校、账号、密码列
                required_cols = ["学校名称/ID", "账号", "密码"]
                for col in required_cols:
                    if col not in headers:
                        return [], f"Excel 文件缺少必需的列：{col}"

                school_idx = headers.index("学校名称/ID")
                account_idx = headers.index("账号")
                password_idx = headers.index("密码")
                remark_idx = headers.index("备注") if "备注" in headers else -1

                for row_idx in range(2, sheet.max_row + 1):
                    row = sheet[row_idx]
                    school = row[school_idx].value
                    account = row[account_idx].value
                    password = row[password_idx].value

                    if not school or not account or not password:
                        continue  # 跳过空行

                    school = str(school).strip()
                    account = str(account).strip()
                    password = str(password).strip()

                    remark = row[remark_idx].value if remark_idx >= 0 else ""
                    remark = str(remark).strip() if remark else ""

                    task = BatchTask(
                        mode="login",
                        school=school,
                        account=account,
                        password=password,
                        remark=remark
                    )
                    tasks.append(task)

            else:
                return [], f"不支持的模式：{mode}"

            workbook.close()

            if errors:
                return tasks, "解析完成，但存在错误：\n" + "\n".join(errors)

            return tasks, ""

        except InvalidFileException:
            return [], "无效的 Excel 文件格式"
        except Exception as e:
            return [], f"解析 Excel 文件时出错：{str(e)}"

    def cancel(self):
        """取消批量处理"""
        self._cancel_event.set()

    def reset_cancel(self):
        """重置取消标志"""
        self._cancel_event.clear()

    def run_batch(self, tasks: List[BatchTask],
                  progress_callback: Optional[Callable[[int, int, str], None]] = None,
                  result_callback: Optional[Callable[[TaskResult], None]] = None) -> BatchResult:
        """
        批量执行任务

        Args:
            tasks: 任务列表
            progress_callback: 进度回调 (current, total, message)
            result_callback: 单个任务结果回调

        Returns:
            BatchResult 批量处理结果
        """
        result = BatchResult()
        result.total = len(tasks)

        for idx, task in enumerate(tasks):
            if self._cancel_event.is_set():
                if progress_callback:
                    progress_callback(idx, result.total, "用户取消操作")
                break

            # 调用进度回调
            if progress_callback:
                msg = f"正在处理第 {idx + 1}/{result.total} 个账号 ({task.account or task.userid})"
                progress_callback(idx, result.total, msg)

            # 处理单个任务
            task_result = self._process_single(task)
            result.details.append(task_result)

            if task_result.success:
                result.success_count += 1
            else:
                result.fail_count += 1

            # 调用结果回调
            if result_callback:
                result_callback(task_result)

        # 最终进度更新
        if progress_callback:
            progress_callback(result.total, result.total, "批量处理完成")

        return result

    def _process_single(self, task: BatchTask) -> TaskResult:
        """
        处理单个任务

        Args:
            task: 批量任务

        Returns:
            TaskResult 任务结果
        """
        import time

        start_time = time.time()

        def log(msg):
            pass

        def check_cancel():
            if self._cancel_event.is_set():
                raise engine.RunCancelled()

        try:
            if task.mode == "userid":
                result = engine.run_by_userid(task.userid, log, check_cancel)
            elif task.mode == "login":
                # 学校名称或 ID 的解析统一由 engine.run_by_login 处理
                result = engine.run_by_login(task.school, task.account, task.password, log, check_cancel)
            else:
                duration = time.time() - start_time
                return TaskResult(task, False, 0, f"不支持的模式：{task.mode}", duration)

            duration = time.time() - start_time
            return TaskResult(task, True, result["score"], "", duration)

        except engine.RunCancelled:
            duration = time.time() - start_time
            return TaskResult(task, False, 0, "用户取消", duration)
        except engine.EngineError as e:
            duration = time.time() - start_time
            return TaskResult(task, False, 0, str(e), duration)
        except Exception as e:
            duration = time.time() - start_time
            return TaskResult(task, False, 0, f"未知错误：{str(e)}", duration)